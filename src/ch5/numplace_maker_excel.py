import numplace_maker
import openpyxl as excel

# Excelファイルを開く --- (*1)
book = excel.load_workbook('numplace_template.xlsx')
sheet = book.worksheets[0]
for y in range(3):
    for x in range(3):
        # 問題を作成 --- (*2)
        data = numplace_maker.shuffle_table([], 0.5)
        for row in range(9):
            for col in range(9):
                # 書き込むセルを計算 --- (*3)
                ix, iy = x*10+col+2, y*10+row+2
                val = data[row][col] if 0 != data[row][col] else ''
                # Excelシートに書き込む --- (*4)
                sheet.cell(iy, ix, val)
                # 空白セルに色を付ける
                if val == '':
                    fill = excel.styles.PatternFill(
                        patternType='solid', fgColor='C0E0C0')
                    sheet.cell(iy, ix).fill = fill
# Excelファイルを保存 --- (*5)
book.save('numplace_data.xlsx')
